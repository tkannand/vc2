"""
Export: Notice-pending voters in WARD 27 who are DMK+ or Neutral (no party).
Sheet 1: Voter-level details
Sheet 2: Summary — ward & booth level stats with party breakdown
"""

import os, sys
sys.path.insert(0, os.path.dirname(__file__))

from dotenv import load_dotenv
load_dotenv()

from backend import storage
from backend.routes_telecaller import _get_alliance

WARD = "WARD 27"

print(f"Loading booths for {WARD}...")
booths = storage.get_booths_for_ward(WARD)
print(f"Found {len(booths)} booths")

rows = []
booth_stats = []

for booth in booths:
    print(f"  Processing {booth}...")
    voters = storage.get_notice_voters_by_booth(WARD, booth)
    statuses = storage.get_all_notice_statuses(WARD, booth)

    booth_num = booth.replace("Booth # ", "").strip() if "Booth #" in booth else booth

    # Per-booth counters
    bs = {
        "booth": booth_num, "total": len(voters),
        "delivered": 0, "pending": 0,
        "delivered_dmk": 0, "delivered_admk": 0, "delivered_ntk": 0,
        "delivered_tvk": 0, "delivered_others": 0, "delivered_neutral": 0,
        "pending_dmk": 0, "pending_neutral": 0,
    }

    for v in voters:
        vid = v.get("RowKey", "")
        status = statuses.get(vid, {}).get("status", "not_delivered")
        party = v.get("party_support", "")
        alliance = _get_alliance(party) if party else ""

        is_delivered = status == "delivered"
        if is_delivered:
            bs["delivered"] += 1
        else:
            bs["pending"] += 1

        # Delivered by party
        if is_delivered:
            if alliance == "DMK+":       bs["delivered_dmk"] += 1
            elif alliance == "ADMK+":    bs["delivered_admk"] += 1
            elif alliance == "NTK":      bs["delivered_ntk"] += 1
            elif alliance == "TVK":      bs["delivered_tvk"] += 1
            elif alliance == "Others":   bs["delivered_others"] += 1
            elif alliance == "":         bs["delivered_neutral"] += 1
        else:
            # Pending DMK+ and Neutral
            if alliance == "DMK+":       bs["pending_dmk"] += 1
            elif alliance == "":         bs["pending_neutral"] += 1

        # Sheet 1: only pending DMK+ or Neutral
        if is_delivered:
            continue
        if alliance not in ("DMK+", ""):
            continue

        # Decrypt phone
        phone = ""
        for field in ("phone_sr_enc", "phone_enc", "whatsapp_enc", "phone3_enc"):
            enc = v.get(field, "")
            if enc:
                p = storage.decrypt_phone(enc)
                if p and len(p) >= 4:
                    phone = p
                    break

        rows.append({
            "SL": v.get("sl", ""),
            "EPIC ID": vid,
            "Booth Number": booth_num,
            "Ward": WARD,
            "Name": v.get("name_ta", "") or v.get("name_en", "") or v.get("name", ""),
            "Street (Tamil)": v.get("section_name_ta", "") or storage.street_key(v),
            "Party Support": party if party else "Neutral",
            "Phone": phone,
        })

    bs["delivered_pct"] = round(bs["delivered"] / bs["total"] * 100, 1) if bs["total"] else 0
    booth_stats.append(bs)

print(f"\nTotal pending DMK+ / Neutral voters: {len(rows)}")

# ── Write Excel ──────────────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
bold_font = Font(name="Calibri", size=10, bold=True)
normal_font = Font(name="Calibri", size=10)
total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
dmk_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def write_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def auto_width(ws, headers, data_rows):
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row in range(2, data_rows + 2):
            val = str(ws.cell(row=row, column=col).value or "")
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_len + 3, 40)


# ── Sheet 1: Pending Voters ─────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Notice Pending"

headers1 = ["SL", "EPIC ID", "Booth Number", "Ward", "Name", "Street (Tamil)", "Party Support", "Phone"]
write_header(ws1, headers1)

for i, r in enumerate(rows, 2):
    for col, key in enumerate(headers1, 1):
        cell = ws1.cell(row=i, column=col, value=r[key])
        cell.font = normal_font
        cell.border = thin_border
        if key == "Party Support" and r[key] != "Neutral":
            cell.fill = dmk_fill

auto_width(ws1, headers1, len(rows))
ws1.freeze_panes = "A2"

# ── Sheet 2: Summary ────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")

headers2 = [
    "Booth", "Total Voters", "Delivered", "Pending", "Delivered %",
    "Del. DMK+", "Del. ADMK+", "Del. NTK", "Del. TVK", "Del. Others", "Del. Neutral",
    "Pending DMK+", "Pending Neutral",
]
write_header(ws2, headers2)

# Ward total row first
ward_total = {k: sum(b[k] for b in booth_stats) for k in [
    "total", "delivered", "pending",
    "delivered_dmk", "delivered_admk", "delivered_ntk",
    "delivered_tvk", "delivered_others", "delivered_neutral",
    "pending_dmk", "pending_neutral",
]}
ward_total["delivered_pct"] = round(ward_total["delivered"] / ward_total["total"] * 100, 1) if ward_total["total"] else 0

row_num = 2
ward_vals = [
    f"{WARD} (Total)", ward_total["total"], ward_total["delivered"], ward_total["pending"],
    f'{ward_total["delivered_pct"]}%',
    ward_total["delivered_dmk"], ward_total["delivered_admk"],
    ward_total["delivered_ntk"], ward_total["delivered_tvk"],
    ward_total["delivered_others"], ward_total["delivered_neutral"],
    ward_total["pending_dmk"], ward_total["pending_neutral"],
]
for col, val in enumerate(ward_vals, 1):
    cell = ws2.cell(row=row_num, column=col, value=val)
    cell.font = bold_font
    cell.fill = total_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

# Booth rows
for bs in booth_stats:
    row_num += 1
    vals = [
        bs["booth"], bs["total"], bs["delivered"], bs["pending"],
        f'{bs["delivered_pct"]}%',
        bs["delivered_dmk"], bs["delivered_admk"],
        bs["delivered_ntk"], bs["delivered_tvk"],
        bs["delivered_others"], bs["delivered_neutral"],
        bs["pending_dmk"], bs["pending_neutral"],
    ]
    for col, val in enumerate(vals, 1):
        cell = ws2.cell(row=row_num, column=col, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

auto_width(ws2, headers2, len(booth_stats) + 1)
ws2.freeze_panes = "A2"

out_path = os.path.join(os.path.dirname(__file__), "notice_pending_ward27.xlsx")
wb.save(out_path)
print(f"Saved to {out_path}")
