import csv
import json
import os
import structlog
import openpyxl
from collections import defaultdict
from backend.config import settings
from backend import storage

logger = structlog.get_logger()


def _decode_tamil(text: str) -> str:
    """Decode Tamil compact encoding (each byte = Tamil Unicode - 0x0B00) to proper Unicode.

    Indian government election CSVs often store Tamil script as single bytes
    where byte value = lower byte of the Tamil Unicode block (U+0B80-U+0BFF).
    ASCII bytes (< 0x80) are passed through unchanged.
    """
    result = []
    for ch in text:
        b = ord(ch)
        if b >= 0x80:
            result.append(chr(0x0B00 + b))
        else:
            result.append(ch)
    return "".join(result).strip()


# ---------------------------------------------------------------------------
# Voter Data — load ONCE from Voter.xlsx
# ---------------------------------------------------------------------------

def rebuild_dashboard_cache() -> dict:
    """Rebuild ward/booth/voter-count cache from the Voters table.

    Run this once after voter data is already loaded to populate the cache.
    Scans the Voters table once — subsequent dashboard loads will be instant.
    """
    from collections import defaultdict
    logger.info("dashboard_cache_rebuild_started")
    table = storage.get_table(storage.table_name("Voters"))
    voter_counts: dict = {}
    seg_counts: dict = {}
    booths_per_ward: dict = defaultdict(set)
    _existing = storage.get_universe_stats()
    # "surveyed_by_party" marker forces recompute when switching from seg_synced to party_support
    need_universe = not _existing or "surveyed_by_party" not in _existing

    # Fields needed for demographics (only fetched if universe stats are missing)
    select_fields = ["PartitionKey", "ward", "booth", "seg_synced", "party_support"]
    if need_universe:
        select_fields += ["gender", "age", "famcode", "section", "section_name"]

    gender_all  = {"M": 0, "F": 0, "O": 0}
    gender_seg  = {"M": 0, "F": 0, "O": 0}
    age_buckets = {"18_25": 0, "26_35": 0, "36_45": 0, "46_60": 0, "61_plus": 0}
    all_famcodes: set      = set()
    surveyed_famcodes: set = set()
    not_surv_famcodes: set = set()
    all_sections: set      = set()
    total_all = total_seg = 0
    surveyed_in_family = not_surv_in_family = 0
    surveyed_ungrouped = not_surv_ungrouped = 0

    for entity in table.query_entities("", select=select_fields):
        ward_raw  = entity.get("ward", "")
        booth_raw = entity.get("booth", "")
        pk        = entity.get("PartitionKey", "")
        if not (ward_raw and booth_raw and pk):
            continue
        voter_counts[pk] = voter_counts.get(pk, 0) + 1
        booths_per_ward[ward_raw].add(booth_raw)
        is_seg = entity.get("seg_synced") == "true"
        if is_seg:
            seg_counts[pk] = seg_counts.get(pk, 0) + 1

        if need_universe:
            total_all += 1
            has_party = bool((entity.get("party_support") or "").strip())
            if has_party: total_seg += 1
            g = (entity.get("gender") or "").strip().upper()
            if g in ("M", "MALE"):
                gender_all["M"] += 1
                if has_party: gender_seg["M"] += 1
            elif g in ("F", "FEMALE"):
                gender_all["F"] += 1
                if has_party: gender_seg["F"] += 1
            else:
                gender_all["O"] += 1
                if has_party: gender_seg["O"] += 1
            try:
                age = int(entity.get("age") or 0)
            except (ValueError, TypeError):
                age = 0
            if age >= 18:
                if age <= 25:   age_buckets["18_25"] += 1
                elif age <= 35: age_buckets["26_35"] += 1
                elif age <= 45: age_buckets["36_45"] += 1
                elif age <= 60: age_buckets["46_60"] += 1
                else:           age_buckets["61_plus"] += 1
            fc = (entity.get("famcode") or "").strip()
            if fc:
                fk = f"{ward_raw}__{booth_raw}__{fc}"
                all_famcodes.add(fk)
                if has_party:
                    surveyed_in_family += 1
                    surveyed_famcodes.add(fk)
                else:
                    not_surv_in_family += 1
                    not_surv_famcodes.add(fk)
            else:
                if has_party: surveyed_ungrouped += 1
                else:         not_surv_ungrouped += 1
            section = (entity.get("section_name") or entity.get("section") or "").strip()
            if section:
                all_sections.add((ward_raw, booth_raw, section))

    wards_list  = sorted(booths_per_ward.keys())
    booths_dict = {w: sorted(list(b)) for w, b in booths_per_ward.items()}

    universe = None
    if need_universe:
        total_booths_count = sum(len(b) for b in booths_per_ward.values())
        universe = {
            "total_voters":       total_all,
            "surveyed_voters":    total_seg,
            "total_families":     len(all_famcodes),
            "surveyed_in_family": surveyed_in_family,
            "surveyed_ungrouped": surveyed_ungrouped,
            "not_surv_in_family": not_surv_in_family,
            "not_surv_ungrouped": not_surv_ungrouped,
            "surveyed_families":  len(surveyed_famcodes),
            "not_surv_families":  len(not_surv_famcodes),
            "ungrouped_voters":   surveyed_ungrouped + not_surv_ungrouped,
            "total_wards":       len(wards_list),
            "total_booths":      total_booths_count,
            "total_streets":     len(all_sections),
            "gender":            gender_all,
            "gender_seg":        gender_seg,
            "age_distribution": [
                {"bucket": "18-25", "count": age_buckets["18_25"]},
                {"bucket": "26-35", "count": age_buckets["26_35"]},
                {"bucket": "36-45", "count": age_buckets["36_45"]},
                {"bucket": "46-60", "count": age_buckets["46_60"]},
                {"bucket": "61+",   "count": age_buckets["61_plus"]},
            ],
            "surveyed_by_party": True,
        }

    storage.store_dashboard_cache(voter_counts, wards_list, booths_dict, seg_counts, universe=universe)
    logger.info("dashboard_cache_rebuild_done", wards=len(wards_list), booths=len(voter_counts),
                universe_stored=bool(universe),
                universe_keys=list(universe.keys()) if universe else [],
                total_voters=universe.get("total_voters") if universe else None,
                ungrouped=universe.get("ungrouped_voters") if universe else None)
    return {"wards": len(wards_list), "booth_partitions": len(voter_counts)}


def _rebuild_booth_meta_from_xlsx() -> None:
    """Read Voter.xlsx to populate booth metadata (name, number, Tamil) in Settings.

    Called once on first restart after this feature was added.
    """
    path = settings.VOTER_DATA_FILE_PATH
    if not os.path.exists(path):
        logger.warning("voter_data_file_not_found_for_booth_meta", path=path)
        return

    booth_meta: dict = {}
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    for row in ws.iter_rows(min_row=2, values_only=True):
        ac_name = str(row[col.get("AC_NAME", 0)] or "").strip()
        booth_num_raw = str(row[col.get("BOOTH", 4)] or "").strip()
        if not ac_name or not booth_num_raw:
            continue
        # Generate "Booth # {number}" format to match existing partition keys
        booth_raw = f"Booth # {int(float(booth_num_raw))}" if booth_num_raw.replace(".", "").isdigit() else f"Booth # {booth_num_raw}"
        ward_raw = ac_name
        if not storage.normalize_key(ward_raw) or not storage.normalize_key(booth_raw):
            continue
        key = (ward_raw, booth_raw)
        if key not in booth_meta:
            booth_meta[key] = {
                "booth_number": booth_num_raw,
                "booth_name": str(row[col.get("BOOTH NAME", 17)] or "").strip(),
                "booth_name_tamil": str(row[col.get("BOOTH NAME - Tamil", 18)] or "").strip(),
            }
    wb.close()

    for (ward_raw, booth_raw), meta in booth_meta.items():
        storage.store_booth_meta(
            ward_raw, booth_raw,
            meta["booth_number"], meta["booth_name"],
            meta["booth_name_tamil"],
        )
    storage.set_setting("booth_meta_stored", "true", "system")
    logger.info("booth_meta_rebuilt", count=len(booth_meta))


def sync_voter_data_once() -> dict:
    """Load Voter.xlsx merged with segment.csv into the Voters table exactly once.

    All columns from both sources are stored on each voter entity.
    Ward/booth come from seg_data (Piv1/Booth) so partition keys match user assignments.
    Guards against re-running with the 'voter_data_loaded' flag in Settings.
    """
    already = storage.get_setting("voter_data_loaded")
    if already == "true":
        _universe = storage.get_universe_stats()
        _needs_rebuild = (
            not storage.get_setting("cached_wards") or
            not _universe or
            "surveyed_by_party" not in _universe
        )
        logger.info("universe_stats_check",
                    exists=bool(_universe),
                    has_detail="surveyed_by_party" in _universe if _universe else False,
                    keys=list(_universe.keys()) if _universe else [])
        if _needs_rebuild:
            logger.info("dashboard_cache_or_universe_missing_rebuilding")
            rebuild_dashboard_cache()
        if not storage.get_setting("booth_meta_stored"):
            logger.info("booth_meta_missing_rebuilding")
            _rebuild_booth_meta_from_xlsx()
        else:
            logger.info("voter_data_already_loaded_skipping")
        return {"synced": 0, "skipped": 0, "status": "already_loaded"}

    voter_path = settings.VOTER_DATA_FILE_PATH
    seg_path   = settings.EXCEL_FILE_PATH
    if not os.path.exists(voter_path):
        logger.warning("voter_data_file_not_found", path=voter_path)
        return {"synced": 0, "skipped": 0, "error": "Voter data file not found"}
    if not os.path.exists(seg_path):
        logger.warning("seg_data_file_not_found", path=seg_path)
        return {"synced": 0, "skipped": 0, "error": "Seg data file not found"}

    # ── Pass 1: build seg_data lookup ────────────────────────────────────────
    # seg_lookup:      voter_id → complete seg row dict
    # booth_num_to_wb: "134"   → ("WARD 12A", "Booth # 134")
    #   Used to assign ward/booth to the ~78K voters not in seg_data
    logger.info("voter_sync_building_seg_lookup")
    seg_lookup: dict    = {}
    booth_num_to_wb: dict = {}

    with open(seg_path, encoding="latin-1", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames and reader.fieldnames[0].startswith("\xff"):
            reader.fieldnames[0] = reader.fieldnames[0].lstrip("\xff")
        for row in reader:
            vid   = (row.get("Voter ID") or "").strip()
            ward  = (row.get("Piv1")     or "").strip()
            booth = (row.get("Booth")    or "").strip()
            if not vid or not ward or not booth:
                continue
            seg_lookup[vid] = row
            bnum = str(int(booth.replace("Booth #", "").strip() or "0"))
            if bnum != "0" and bnum not in booth_num_to_wb:
                booth_num_to_wb[bnum] = (ward, booth)

    logger.info("voter_sync_seg_lookup_built",
                seg_rows=len(seg_lookup), booths_mapped=len(booth_num_to_wb))

    # ── Pass 2: read voter_data (xlsx), merge ALL columns into one entity ────
    logger.info("voter_data_sync_started", path=voter_path)
    synced = 0
    skipped = 0
    errors  = 0
    batch   = []
    seen_epics: set = set()  # deduplicate — xlsx may have duplicate EPICs
    skip_reasons = {"no_epic": 0, "in_seg": 0, "booth_fallback": 0,
                    "no_booth_match": 0, "duplicate": 0, "deleted": 0, "row_error": 0}

    wb = openpyxl.load_workbook(voter_path, read_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    for row_tuple in ws.iter_rows(min_row=2, values_only=True):
        voter_id = ""
        try:
            voter_id = str(row_tuple[col.get("EPIC", 6)] or "").strip()
            if not voter_id or voter_id == "None":
                skip_reasons["no_epic"] += 1
                skipped += 1
                continue
            if voter_id in seen_epics:
                skip_reasons["duplicate"] += 1
                skipped += 1
                continue
            # Skip deleted voters
            is_deleted = str(row_tuple[col.get("IS_DELETED", 16)] or "FALSE").strip().upper() == "TRUE"
            if is_deleted:
                skip_reasons["deleted"] += 1
                skipped += 1
                continue
            seen_epics.add(voter_id)

            seg = seg_lookup.get(voter_id, {})

            # Ward / booth -- from seg_data (Piv1 / Booth)
            ward  = (seg.get("Piv1")  or "").strip()
            booth = (seg.get("Booth") or "").strip()
            if ward and booth:
                skip_reasons["in_seg"] += 1  # will sync
            else:
                # Voter not in seg -- infer ward from booth number
                raw = str(row_tuple[col.get("BOOTH", 4)] or "").strip()
                bnum = str(int(float(raw))) if raw.replace(".", "").isdigit() else raw
                if bnum in booth_num_to_wb:
                    ward, booth = booth_num_to_wb[bnum]
                    skip_reasons["booth_fallback"] += 1  # will sync
                else:
                    skip_reasons["no_booth_match"] += 1
                    skipped += 1
                    continue

            # ── Voter_Data columns (xlsx -- already Unicode, no decoding) ──
            name_en          = str(row_tuple[col.get("NAME_T", 8)] or "").strip().rstrip(" -").strip()
            name_ta          = str(row_tuple[col.get("NAME_EN", 9)] or "").strip().rstrip(" -").strip()
            relation_name_en = str(row_tuple[col.get("R_Name", 11)] or "").strip().rstrip(" -").strip()
            relation_name_ta = str(row_tuple[col.get("REL_EN", 12)] or "").strip().rstrip(" -").strip()
            relation_type    = str(row_tuple[col.get("R_TYPE", 10)] or "").strip()
            booth_num_raw    = str(row_tuple[col.get("BOOTH", 4)] or "").strip()
            # Generate "Booth # {number}" format from raw booth number
            booth_number     = str(int(float(booth_num_raw))) if booth_num_raw.replace(".", "").isdigit() else booth_num_raw
            booth_display    = f"Booth # {booth_number}"
            booth_name       = str(row_tuple[col.get("BOOTH NAME", 17)] or "").strip()
            booth_name_ta    = str(row_tuple[col.get("BOOTH NAME - Tamil", 18)] or "").strip()
            section_num      = str(row_tuple[col.get("SECTION", 1)] or "").strip()
            section_name     = str(row_tuple[col.get("SECTION_NAME", 2)] or "").strip()
            # Tamil section name -- strip "{number}::" prefix, skip bad values
            section_name_ta_raw = str(row_tuple[col.get("SECTION_NAME - Tamil", 3)] or "").strip()
            section_name_ta  = ""
            if section_name_ta_raw and section_name_ta_raw not in ("#N/A", "ADD_LIST::ADD_LIST"):
                if "::" in section_name_ta_raw:
                    section_name_ta = section_name_ta_raw.split("::", 1)[1].strip()
                else:
                    section_name_ta = section_name_ta_raw
            ac_name          = str(row_tuple[col.get("AC_NAME", 0)] or "").strip()
            house_vd         = str(row_tuple[col.get("HOUSE", 7)] or "").strip()
            sl               = str(row_tuple[col.get("SL", 5)] or "").strip()
            is_deleted       = str(row_tuple[col.get("IS_DELETED", 16)] or "FALSE").strip().upper() == "TRUE"

            age_vd = 0
            try:
                age_vd = int(float(row_tuple[col.get("AGE", 13)] or 0))
            except (ValueError, TypeError):
                pass
            gender_vd = str(row_tuple[col.get("GENDER", 14)] or "").strip()

            # ── Seg_Data columns (all plain ASCII/English) ───────────────
            def _s(v): return (v or "").strip()
            piv0          = _s(seg.get("Piv0"))
            piv2          = _s(seg.get("Piv2"))
            name_seg      = _s(seg.get("Name"))
            rel_name_seg  = _s(seg.get("Relation Name"))
            relationship  = _s(seg.get("Relationship"))
            house         = storage.fix_excel_date(_s(seg.get("House")) or house_vd)
            house2        = storage.fix_excel_date(_s(seg.get("House2")))
            famcode       = _s(seg.get("Famcode"))
            is_head       = _s(seg.get("Is Head of Household?")) or "No"
            party_support = _s(seg.get("Party Support"))
            phys_disabled = _s(seg.get("Physically Disabled"))
            religion      = _s(seg.get("Religion"))
            caste         = _s(seg.get("Caste"))
            education     = _s(seg.get("Education"))
            occupation    = _s(seg.get("Occupation"))
            econ_status   = _s(seg.get("Economic Status"))
            outside_voter = _s(seg.get("Outside Voter"))
            phone_sr      = _s(seg.get("Phone (SR)"))
            whatsapp      = _s(seg.get("WhatsApp"))
            phone         = _s(seg.get("Phone"))
            phone3        = _s(seg.get("Phone3"))
            ration_card   = _s(seg.get("Ration Card"))
            section_seg   = _s(seg.get("Section"))

            age    = age_vd
            try:
                if seg.get("Age"):
                    age = int(float(seg["Age"]))
            except (ValueError, TypeError):
                pass
            gender = (seg.get("Gender") or gender_vd).strip()

            voter = {
                "voter_id":           voter_id,
                # Names
                "name":               name_en,
                "name_en":            name_en,
                "name_ta":            name_ta,
                "name_seg":           name_seg,
                # Relations
                "relation_name":      relation_name_en,
                "relation_name_ta":   relation_name_ta,
                "relation_name_seg":  rel_name_seg,
                "relationship":       relation_type or relationship,
                # Demographics
                "age":                age,
                "gender":             gender,
                # Address
                "house":              house,
                "house2":             house2,
                # Ward / Booth -- from seg_data (Piv1 / Booth)
                "ward":               ward,        # "WARD 12A"
                "booth":              booth,       # "Booth # 134"
                "ac":                 ac_name,
                "piv0":               piv0,        # "AC_191"
                "piv2":               piv2,        # "Booth #134"
                # Booth metadata -- from voter_data
                "booth_number":       booth_number,   # "240"
                "booth_display":      booth_display,  # "Booth # 240"
                "booth_name":         booth_name,
                "booth_name_tamil":   booth_name_ta,
                # Section / street
                "section":            section_name or section_seg,
                "section_num":        section_num,
                "section_name":       section_name,
                "section_name_ta":    section_name_ta,
                # Seg enrichment
                "famcode":            famcode,
                "is_head":            is_head,
                "party_support":      party_support,
                "physically_disabled": phys_disabled,
                "religion":           religion,
                "caste":              caste,
                "education":          education,
                "occupation":         occupation,
                "economic_status":    econ_status,
                "outside_voter":      outside_voter,
                # Phones (encrypted by _build_voter_entity)
                "phone_sr":           phone_sr,
                "whatsapp":           whatsapp,
                "phone":              phone,
                "phone3":             phone3,
                # Misc
                "ration_card":        ration_card,
                "sl":                 sl,
                "is_deleted":         str(is_deleted),
                "seg_synced":         "true" if seg else "false",
            }

            batch.append(voter)
            synced += 1

        except Exception as e:
            errors += 1
            skip_reasons["row_error"] += 1
            logger.error("voter_data_row_error", voter_id=voter_id, error=str(e)[:300])

    wb.close()

    logger.info("voter_data_batch_uploading", count=len(batch),
                in_seg=skip_reasons["in_seg"], booth_fallback=skip_reasons["booth_fallback"],
                no_booth_match=skip_reasons["no_booth_match"],
                no_epic=skip_reasons["no_epic"], duplicates=skip_reasons["duplicate"],
                deleted=skip_reasons["deleted"],
                row_errors=skip_reasons["row_error"])
    storage.batch_upsert_voters(batch)

    # ── Build dashboard cache + booth metadata in one pass ───────────────────
    voter_counts: dict    = {}  # all voters — used by notice
    seg_counts: dict      = {}  # seg_synced only — used by calling dashboard
    booths_per_ward: dict = defaultdict(set)
    booth_meta: dict      = {}
    for v in batch:
        ward_raw  = v.get("ward", "")
        booth_raw = v.get("booth", "")
        if ward_raw and booth_raw:
            pk = f"{storage.normalize_key(ward_raw)}__{storage.normalize_key(booth_raw)}"
            voter_counts[pk] = voter_counts.get(pk, 0) + 1
            if v.get("seg_synced") == "true":
                seg_counts[pk] = seg_counts.get(pk, 0) + 1
            booths_per_ward[ward_raw].add(booth_raw)
            key = (ward_raw, booth_raw)
            if key not in booth_meta and (v.get("booth_name") or v.get("booth_number")):
                booth_meta[key] = {
                    "booth_number":      v.get("booth_number", ""),
                    "booth_name":        v.get("booth_name", ""),
                    "booth_name_tamil":  v.get("booth_name_tamil", ""),
                }

    wards_list  = sorted(booths_per_ward.keys())
    booths_dict = {w: sorted(list(b)) for w, b in booths_per_ward.items()}

    # ── Compute demographic / universe stats (in-memory, no extra scan) ──────
    gender_all  = {"M": 0, "F": 0, "O": 0}
    gender_seg  = {"M": 0, "F": 0, "O": 0}
    age_buckets = {"18_25": 0, "26_35": 0, "36_45": 0, "46_60": 0, "61_plus": 0}
    all_famcodes: set      = set()
    surveyed_famcodes: set = set()   # unique famcodes among surveyed voters
    not_surv_famcodes: set = set()   # unique famcodes among not-surveyed voters
    all_sections: set      = set()
    total_surveyed       = 0
    surveyed_in_family   = 0   # surveyed + has famcode
    not_surv_in_family   = 0   # not surveyed + has famcode
    surveyed_ungrouped   = 0   # surveyed + no famcode
    not_surv_ungrouped   = 0   # not surveyed + no famcode

    for v in batch:
        has_party = bool((v.get("party_support") or "").strip())
        if has_party:
            total_surveyed += 1

        # Gender
        g = (v.get("gender") or "").strip().upper()
        if g in ("M", "MALE"):
            gender_all["M"] += 1
            if has_party:
                gender_seg["M"] += 1
        elif g in ("F", "FEMALE"):
            gender_all["F"] += 1
            if has_party:
                gender_seg["F"] += 1
        else:
            gender_all["O"] += 1
            if has_party:
                gender_seg["O"] += 1

        # Age
        try:
            age = int(v.get("age") or 0)
        except (ValueError, TypeError):
            age = 0
        if age >= 18:
            if age <= 25:
                age_buckets["18_25"] += 1
            elif age <= 35:
                age_buckets["26_35"] += 1
            elif age <= 45:
                age_buckets["36_45"] += 1
            elif age <= 60:
                age_buckets["46_60"] += 1
            else:
                age_buckets["61_plus"] += 1

        # Families — split by surveyed (has party support) status
        fc = (v.get("famcode") or "").strip()
        if fc:
            fk = f"{v.get('ward', '')}__{v.get('booth', '')}__{fc}"
            all_famcodes.add(fk)
            if has_party:
                surveyed_in_family += 1
                surveyed_famcodes.add(fk)
            else:
                not_surv_in_family += 1
                not_surv_famcodes.add(fk)
        else:
            if has_party:
                surveyed_ungrouped += 1
            else:
                not_surv_ungrouped += 1

        # Sections / streets (unique per ward+booth+section)
        section = (v.get("section") or v.get("section_name") or "").strip()
        if section and v.get("ward") and v.get("booth"):
            all_sections.add((v["ward"], v["booth"], section))

    total_booths_count = sum(len(b) for b in booths_per_ward.values())
    universe = {
        "total_voters":         synced,
        "surveyed_voters":      total_surveyed,
        "total_families":       len(all_famcodes),
        "surveyed_in_family":   surveyed_in_family,
        "surveyed_ungrouped":   surveyed_ungrouped,
        "not_surv_in_family":   not_surv_in_family,
        "not_surv_ungrouped":   not_surv_ungrouped,
        "surveyed_families":    len(surveyed_famcodes),
        "not_surv_families":    len(not_surv_famcodes),
        "ungrouped_voters":     surveyed_ungrouped + not_surv_ungrouped,
        "total_wards":       len(wards_list),
        "total_booths":      total_booths_count,
        "total_streets":     len(all_sections),
        "gender":            gender_all,
        "gender_seg":        gender_seg,
        "age_distribution": [
            {"bucket": "18-25", "count": age_buckets["18_25"]},
            {"bucket": "26-35", "count": age_buckets["26_35"]},
            {"bucket": "36-45", "count": age_buckets["36_45"]},
            {"bucket": "46-60", "count": age_buckets["46_60"]},
            {"bucket": "61+",   "count": age_buckets["61_plus"]},
        ],
        "surveyed_by_party": True,
    }
    storage.store_dashboard_cache(voter_counts, wards_list, booths_dict, seg_counts, universe=universe)

    for (ward_raw, booth_raw), meta in booth_meta.items():
        storage.store_booth_meta(
            ward_raw, booth_raw,
            meta["booth_number"], meta["booth_name"],
            meta["booth_name_tamil"],
        )
    logger.info("booth_meta_stored", count=len(booth_meta))
    storage.set_setting("booth_meta_stored", "true", "system")

    storage.set_setting("voter_data_loaded", "true", "system")
    logger.info("voter_data_sync_completed", synced=synced, skipped=skipped, errors=errors)
    return {"synced": synced, "skipped": skipped, "errors": errors}


# ---------------------------------------------------------------------------
# Seg Data — incremental sync from segment.csv
# ---------------------------------------------------------------------------

def sync_seg_data_incremental() -> dict:
    """Incrementally sync segment.csv on every restart.

    Fetches the set of voter IDs already merged from seg data (via the
    seg_synced flag on each voter entity). Only processes rows whose
    voter_id is not in that set — completely order/sort independent.

    Seg fields are MERGED into existing voter records — Tamil names and
    other voter-data fields are never overwritten.
    """
    path = settings.EXCEL_FILE_PATH
    if not os.path.exists(path):
        logger.warning("seg_data_file_not_found", path=path)
        return {"synced": 0, "skipped": 0, "error": "File not found"}

    # Load already-synced voter IDs from Azure (projection query — fast)
    already_synced = storage.get_seg_synced_voter_ids()

    logger.info("seg_sync_started", path=path, already_synced=len(already_synced))
    batch = []  # list of (voter_id, ward, booth, seg) tuples
    skipped = 0
    errors = 0
    # Collect every (ward, booth) pair from the full CSV — used to build the
    # notice ward/booth cache regardless of whether the voter is already synced
    seg_wards_set: set = set()
    seg_booths_per_ward: dict = defaultdict(set)

    with open(path, encoding="latin-1", newline="") as f:
        reader = csv.DictReader(f)
        # Strip leading 0xFF BOM byte that corrupts the first column name
        if reader.fieldnames and reader.fieldnames[0].startswith("\xff"):
            reader.fieldnames[0] = reader.fieldnames[0].lstrip("\xff")
        for row in reader:
            try:
                voter_id = (row.get("Voter ID") or "").strip()
                if not voter_id or voter_id == "None":
                    skipped += 1
                    continue

                ward = (row.get("Piv1") or "").strip()
                booth = (row.get("Booth") or "").strip()
                if not ward or not booth:
                    skipped += 1
                    continue

                # Always track ward/booth for notice cache (even already-synced rows)
                seg_wards_set.add(ward)
                seg_booths_per_ward[ward].add(booth)

                if voter_id in already_synced:
                    skipped += 1
                    continue

                age = 0
                try:
                    age = int(float(row.get("Age") or 0))
                except (ValueError, TypeError):
                    age = 0

                seg = {
                    "name": (row.get("Name") or "").strip(),
                    "relation_name": (row.get("Relation Name") or "").strip(),
                    "relationship": (row.get("Relationship") or "").strip(),
                    "age": age,
                    "gender": (row.get("Gender") or "").strip(),
                    "house": (row.get("House") or "").strip(),
                    "house2": (row.get("House2") or "").strip(),
                    "famcode": (row.get("Famcode") or "").strip(),
                    "is_head": (row.get("Is Head of Household?") or "No").strip(),
                    "party_support": (row.get("Party Support") or "").strip(),
                    "physically_disabled": (row.get("Physically Disabled") or "").strip(),
                    "religion": (row.get("Religion") or "").strip(),
                    "caste": (row.get("Caste") or "").strip(),
                    "education": (row.get("Education") or "").strip(),
                    "occupation": (row.get("Occupation") or "").strip(),
                    "economic_status": (row.get("Economic Status") or "").strip(),
                    "outside_voter": (row.get("Outside Voter") or "").strip(),
                    "phone_sr": (row.get("Phone (SR)") or "").strip(),
                    "whatsapp": (row.get("WhatsApp") or "").strip(),
                    "phone": (row.get("Phone") or "").strip(),
                    "phone3": (row.get("Phone3") or "").strip(),
                    "booth": booth,
                    "ward": ward,
                    "ac": (row.get("Piv0") or "").strip(),
                    "booth_full": (row.get("Piv2") or "").strip(),
                    "ration_card": (row.get("Ration Card") or "").strip(),
                    "section": (row.get("Section") or "").strip(),
                }

                batch.append((voter_id, ward, booth, seg))

            except Exception as e:
                errors += 1
                logger.error("seg_row_sync_error", error=str(e))

    synced = 0
    if batch:
        logger.info("seg_batch_uploading", count=len(batch))
        synced = storage.batch_merge_voter_seg_data(batch)

    # Always rebuild the seg ward/booth cache from the full CSV — this drives
    # the notice system's ward enumeration (distinct from voter_data AC ward)
    if seg_wards_set:
        storage.set_setting("cached_seg_wards", json.dumps(sorted(seg_wards_set)), "system")
        for w, booths in seg_booths_per_ward.items():
            storage.set_setting(
                f"cached_seg_booths_{storage.normalize_key(w)}",
                json.dumps(sorted(list(booths))),
                "system",
            )
        logger.info("seg_ward_cache_stored", wards=len(seg_wards_set))

    logger.info("seg_sync_completed", synced=synced, skipped=skipped, errors=errors)
    return {"synced": synced, "skipped": skipped, "errors": errors}
