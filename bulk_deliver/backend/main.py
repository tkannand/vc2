import os
import json
import uuid
import tempfile
import structlog
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from contextlib import asynccontextmanager
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from bulk_deliver.backend import storage
from bulk_deliver.backend.processor import process_deliveries

logger = structlog.get_logger()

# ── In-memory batch store (single-user admin tool) ──────────────────────────
_batches = {}

# ── Temp directory for report files ─────────────────────────────────────────
_REPORT_DIR = os.path.join(tempfile.gettempdir(), "bulk_deliver_reports")
os.makedirs(_REPORT_DIR, exist_ok=True)


@asynccontextmanager
async def lifespan(app: FastAPI):
    storage.init_audit_table()
    logger.info("bulk_deliver_app_started", report_dir=_REPORT_DIR)
    yield
    logger.info("bulk_deliver_app_stopped")


app = FastAPI(title="Bulk Delivery Tool", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── Upload ──────────────────────────────────────────────────────────────────

@app.post("/api/upload")
async def upload_excel(file: UploadFile = File(...), operator_name: str = Form(...)):
    if not operator_name or not operator_name.strip():
        raise HTTPException(400, "Operator name is required")

    fname = (file.filename or "").lower()
    if not fname.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only Excel files (.xlsx) are supported")

    batch_id = uuid.uuid4().hex[:12]
    temp_path = os.path.join(_REPORT_DIR, f"{batch_id}_upload.xlsx")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(400, "File too large (max 10 MB)")

    with open(temp_path, "wb") as f:
        f.write(content)

    # Parse EPICs
    try:
        wb = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Cannot read Excel file: {exc}")

    ws = wb.active

    # Find EPIC column by header
    epic_col = None
    first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    if first_row:
        for col_idx, cell in enumerate(first_row[0]):
            val = str(cell.value or "").strip().lower()
            if any(kw in val for kw in ["epic", "voter_id", "voterid", "voter id"]):
                epic_col = col_idx
                break

    if epic_col is None:
        epic_col = 0  # default to first column

    epics = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=False):
        if epic_col >= len(row):
            continue
        val = row[epic_col].value
        if val:
            epic = str(val).strip().upper()
            if epic and epic not in seen:
                epics.append(epic)
                seen.add(epic)

    wb.close()

    if not epics:
        raise HTTPException(400, "No EPIC numbers found in the file")

    _batches[batch_id] = {
        "epics": epics,
        "operator_name": operator_name.strip(),
        "source_filename": file.filename,
        "scheme_type": None,
        "scheme_id": None,
        "scheme_name": None,
        "green_voters": [],
    }

    logger.info("excel_uploaded", batch_id=batch_id, epic_count=len(epics),
                filename=file.filename, operator=operator_name.strip())

    return {
        "batch_id": batch_id,
        "epic_count": len(epics),
        "epics_preview": epics[:20],
    }


# ── Schemes ─────────────────────────────────────────────────────────────────

@app.get("/api/schemes")
async def list_schemes():
    schemes = []
    if storage.get_notice_enabled():
        schemes.append({"id": "notice", "name": "Notice", "type": "individual"})
    if storage.get_coupon_enabled():
        schemes.append({"id": "coupon", "name": "Coupon", "type": "family"})
    for cs in storage.get_custom_schemes():
        schemes.append({"id": cs["id"], "name": cs["name"], "type": cs["type"]})
    return {"schemes": schemes}


# ── Diagnostics ─────────────────────────────────────────────────────────────

@app.post("/api/diagnose")
async def run_diagnostics(body: dict):
    batch_id = body.get("batch_id")
    scheme_id = body.get("scheme_id", "")
    scheme_name = body.get("scheme_name", "")

    if batch_id not in _batches:
        raise HTTPException(400, "Invalid batch_id - please re-upload the file")

    batch = _batches[batch_id]

    # Determine scheme_type from scheme_id
    if scheme_id == "notice":
        scheme_type = "notice"
    elif scheme_id == "coupon":
        scheme_type = "coupon"
    else:
        scheme_type = "custom"

    batch["scheme_type"] = scheme_type
    batch["scheme_id"] = scheme_id
    batch["scheme_name"] = scheme_name

    epics = batch["epics"]

    # Load all voters from Azure
    all_voters = storage.load_all_voters()

    green = []
    yellow = []
    red = []

    # Classify each EPIC
    found_voters = {}
    for epic in epics:
        if epic in all_voters:
            found_voters[epic] = all_voters[epic]
        else:
            yellow.append({"voter_id": epic})

    # Group found voters by ward/booth for efficient status lookups
    by_partition = defaultdict(list)
    for vid, voter in found_voters.items():
        key = (voter["ward"], voter["booth"])
        by_partition[key].append(voter)

    # Check delivery status per partition
    for (ward, booth), voters in by_partition.items():
        statuses = storage.get_delivery_statuses(scheme_type, scheme_id, ward, booth)
        for voter in voters:
            vid = voter["voter_id"]
            if vid in statuses and statuses[vid].get("status") == "delivered":
                st = statuses[vid]
                red.append({
                    "voter_id": vid,
                    "name": voter.get("name", ""),
                    "ward": voter.get("ward", ""),
                    "booth": voter.get("booth", ""),
                    "sl": voter.get("sl", ""),
                    "delivered_by_name": st.get("delivered_by_name", ""),
                    "delivered_at": st.get("updated_at", ""),
                })
            else:
                green.append(voter)

    batch["green_voters"] = green

    # Generate downloadable report for red + yellow
    report_file = None
    if red or yellow:
        report_file = f"{batch_id}_diagnostics.xlsx"
        report_path = os.path.join(_REPORT_DIR, report_file)
        _generate_diagnostics_report(report_path, red, yellow)

    logger.info("diagnostics_complete", batch_id=batch_id,
                green=len(green), red=len(red), yellow=len(yellow))

    return {
        "green_count": len(green),
        "red_count": len(red),
        "yellow_count": len(yellow),
        "red": red,
        "yellow": yellow,
        "report_file": report_file,
    }


def _generate_diagnostics_report(path: str, red: list, yellow: list):
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill_red = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    header_fill_yellow = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Already Delivered sheet
    ws1 = wb.active
    ws1.title = "Already Delivered"
    headers_red = ["EPIC", "Name", "Ward", "Booth", "SL No", "Delivered By", "Delivered At"]
    ws1.append(headers_red)
    for i, cell in enumerate(ws1[1], 1):
        cell.font = header_font
        cell.fill = header_fill_red
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for r in red:
        ws1.append([
            r["voter_id"], r.get("name", ""), r.get("ward", ""),
            r.get("booth", ""), r.get("sl", ""),
            r.get("delivered_by_name", ""), r.get("delivered_at", ""),
        ])
    for col in ws1.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Not Found sheet
    ws2 = wb.create_sheet("Not Found in Database")
    headers_yellow = ["EPIC"]
    ws2.append(headers_yellow)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill_yellow
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for y in yellow:
        ws2.append([y["voter_id"]])
    ws2.column_dimensions["A"].width = 20

    wb.save(path)


# ── Download ────────────────────────────────────────────────────────────────

@app.get("/api/download/{filename}")
async def download_report(filename: str):
    # Sanitize filename to prevent path traversal
    safe_name = os.path.basename(filename)
    path = os.path.join(_REPORT_DIR, safe_name)
    if not os.path.exists(path):
        raise HTTPException(404, "Report file not found")
    return FileResponse(
        path,
        filename=safe_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── Process (SSE) ──────────────────────────────────────────────────────────

@app.get("/api/process")
async def process_batch(batch_id: str):
    if batch_id not in _batches:
        raise HTTPException(400, "Invalid batch_id")

    batch = _batches[batch_id]
    voters = batch.get("green_voters", [])

    if not voters:
        raise HTTPException(400, "No voters to process. Run diagnostics first.")

    async def event_stream():
        async for event in process_deliveries(
            voters=voters,
            scheme_type=batch["scheme_type"],
            scheme_id=batch["scheme_id"],
            scheme_name=batch["scheme_name"],
            operator_name=batch["operator_name"],
            batch_id=batch_id,
            source_filename=batch["source_filename"],
        ):
            # If processing is complete and there are failures, generate report
            if event["type"] == "complete" and event.get("failures"):
                fail_file = f"{batch_id}_failures.xlsx"
                fail_path = os.path.join(_REPORT_DIR, fail_file)
                _generate_failure_report(fail_path, event["failures"])
                event["failure_report"] = fail_file

            yield f"data: {json.dumps(event)}\n\n"

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


def _generate_failure_report(path: str, failures: list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Failed Deliveries"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")

    headers = ["EPIC", "Name", "Ward", "Booth", "SL No", "Error"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for f in failures:
        ws.append([
            f.get("voter_id", ""), f.get("name", ""), f.get("ward", ""),
            f.get("booth", ""), f.get("sl", ""), f.get("error", ""),
        ])
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(path)


# ── Serve frontend ──────────────────────────────────────────────────────────

_FRONTEND_DIR = os.path.join(os.path.dirname(__file__), "..", "frontend")


@app.get("/")
async def serve_index():
    return FileResponse(os.path.join(_FRONTEND_DIR, "index.html"))


app.mount("/css", StaticFiles(directory=os.path.join(_FRONTEND_DIR, "css")), name="css")
app.mount("/js", StaticFiles(directory=os.path.join(_FRONTEND_DIR, "js")), name="js")
